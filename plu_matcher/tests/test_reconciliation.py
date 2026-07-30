import json

from openpyxl import Workbook, load_workbook

from src.models import MenuNode
from src.reconciliation import (
    MenuIndexes,
    STATUS_AMBIGUOUS,
    STATUS_ID_UPDATED,
    STATUS_MATCH,
    STATUS_MISMATCH,
    STATUS_MULTIPLE_PATHS,
    STATUS_NOT_FOUND,
    STATUS_NOT_IN_CONTEXT,
    decide_reconciliation,
    reconcile_workbook,
)


def make_node(
    *,
    item_id: int,
    plu: str,
    title: str,
    path: str,
) -> MenuNode:
    return MenuNode(
        item_master_id=item_id,
        title=title,
        plu=plu,
        item_path_key=path,
        parent_path_key=None,
        item_type=1,
        deleted=False,
    )


def make_indexes(*nodes: MenuNode) -> MenuIndexes:
    by_plu: dict[str, list[MenuNode]] = {}
    by_item_id: dict[int, list[MenuNode]] = {}
    for node in nodes:
        by_plu.setdefault(str(node.plu), []).append(node)
        by_item_id.setdefault(node.item_master_id, []).append(node)
    return MenuIndexes(by_plu=by_plu, by_item_id=by_item_id)


def test_unique_item_id_is_safe_to_backfill():
    node = make_node(
        item_id=100,
        plu="900001",
        title="Milk",
        path="1-100",
    )

    decision = decide_reconciliation(
        "900001",
        None,
        "",
        make_indexes(node),
    )

    assert decision.status == STATUS_MATCH
    assert decision.proposed_item_id == 100
    assert decision.apply_mapping is True


def test_one_item_id_on_multiple_paths_is_safe_identity():
    nodes = (
        make_node(
            item_id=100,
            plu="900001",
            title="Milk",
            path="1-100",
        ),
        make_node(
            item_id=100,
            plu="900001",
            title="Milk",
            path="2-100",
        ),
    )

    decision = decide_reconciliation(
        "900001",
        None,
        "",
        make_indexes(*nodes),
    )

    assert decision.status == STATUS_MULTIPLE_PATHS
    assert decision.proposed_item_id == 100
    assert decision.path_count == 2
    assert decision.apply_mapping is True


def test_multiple_item_ids_are_ambiguous_and_not_backfilled():
    nodes = (
        make_node(
            item_id=100,
            plu="900001",
            title="Milk A",
            path="1-100",
        ),
        make_node(
            item_id=200,
            plu="900001",
            title="Milk B",
            path="1-200",
        ),
    )

    decision = decide_reconciliation(
        "900001",
        None,
        "",
        make_indexes(*nodes),
    )

    assert decision.status == STATUS_AMBIGUOUS
    assert decision.proposed_item_id is None
    assert decision.apply_mapping is False
    assert "100:" in decision.candidates
    assert "200:" in decision.candidates


def test_unique_plu_replaces_existing_different_item_id():
    node = make_node(
        item_id=200,
        plu="900001",
        title="Current Milk",
        path="1-200",
    )

    decision = decide_reconciliation(
        "900001",
        100,
        "Existing Milk",
        make_indexes(node),
    )

    assert decision.status == STATUS_ID_UPDATED
    assert decision.proposed_item_id == 200
    assert decision.proposed_item_name == "Current Milk"
    assert decision.apply_mapping is True
    assert "previous QU item 100" in decision.notes


def test_multiple_item_ids_do_not_replace_existing_mapping():
    nodes = (
        make_node(
            item_id=200,
            plu="900001",
            title="Current Milk A",
            path="1-200",
        ),
        make_node(
            item_id=300,
            plu="900001",
            title="Current Milk B",
            path="1-300",
        ),
    )

    decision = decide_reconciliation(
        "900001",
        100,
        "Existing Milk",
        make_indexes(*nodes),
    )

    assert decision.status == STATUS_AMBIGUOUS
    assert decision.proposed_item_id is None
    assert decision.apply_mapping is False


def test_existing_item_with_different_plu_requires_review():
    node = make_node(
        item_id=100,
        plu="900002",
        title="Milk",
        path="1-100",
    )

    decision = decide_reconciliation(
        "900001",
        100,
        "Milk",
        make_indexes(node),
    )

    assert decision.status == STATUS_MISMATCH
    assert decision.apply_mapping is False
    assert "900002" in decision.notes


def test_unmapped_plu_not_in_qu_is_not_found():
    decision = decide_reconciliation(
        "900001",
        None,
        "",
        make_indexes(),
    )

    assert decision.status == STATUS_NOT_FOUND
    assert decision.apply_mapping is False


def test_existing_item_absent_from_context_is_preserved():
    decision = decide_reconciliation(
        "900001",
        100,
        "Milk",
        make_indexes(),
    )

    assert decision.status == STATUS_NOT_IN_CONTEXT
    assert decision.apply_mapping is False


def test_workbook_audit_skips_unmigrated_rows_and_records_provenance(
    tmp_path,
):
    source_path = tmp_path / "source.xlsx"
    menu_path = tmp_path / "menu.json"
    output_path = tmp_path / "output.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Aloha > Qu Item Migration"
    for _ in range(7):
        sheet.append([])
    sheet.append(
        [
            "Number",
            "Long name",
            "Migrated to Qu?",
            "Qu Item ID",
            "Qu Item Name",
        ]
    )
    sheet.append([900001, "Milk", "Yes", 100, "Milk"])
    sheet.append([900002, "Tea", "No", None, None])
    workbook.save(source_path)
    workbook.close()

    menu = {
        "menuSnapshotId": "snapshot-1",
        "context": {
            "locationId": 11934,
            "orderChannelId": 4685,
            "orderTypeId": 4723,
        },
        "children": [
            {
                "itemMasterId": 100,
                "title": "Milk",
                "itemPathKey": "1-100",
                "displayAttribute": {"plu": "900001"},
                "children": [],
            },
            {
                "itemMasterId": 200,
                "title": "Tea",
                "itemPathKey": "1-200",
                "displayAttribute": {"plu": "900002"},
                "children": [],
            },
        ],
    }
    menu_path.write_text(json.dumps(menu), encoding="utf-8")

    counts = reconcile_workbook(
        source_path,
        menu_path,
        output_path,
        provenance={
            "generation_time": "generation-1",
            "last_checked_at": "2026-07-30T12:00:00+00:00",
        },
    )

    assert sum(counts.values()) == 1
    result = load_workbook(output_path, data_only=True)
    mapping = result["Aloha > Qu Item Migration"]
    headers = {
        cell.value: cell.column for cell in mapping[8] if cell.value
    }
    assert (
        mapping.cell(9, headers["Reconciliation Status"]).value
        == STATUS_MATCH.value
    )
    assert (
        mapping.cell(10, headers["Reconciliation Status"]).value
        is None
    )

    summary = result["Reconciliation Summary"]
    summary_values = {
        summary.cell(row, 1).value: summary.cell(row, 2).value
        for row in range(1, summary.max_row + 1)
    }
    assert summary_values["Menu snapshot ID"] == "snapshot-1"
    assert summary_values["Menu generation time"] == "generation-1"
    assert summary_values["Location ID"] == 11934
    result.close()


def test_workbook_can_explicitly_include_unmigrated_rows(tmp_path):
    source_path = tmp_path / "source.xlsx"
    menu_path = tmp_path / "menu.json"
    output_path = tmp_path / "output.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Aloha > Qu Item Migration"
    for _ in range(7):
        sheet.append([])
    sheet.append(
        [
            "Number",
            "Long name",
            "Migrated to Qu?",
            "Qu Item ID",
            "Qu Item Name",
        ]
    )
    sheet.append([900002, "Tea", "No", None, None])
    workbook.save(source_path)
    workbook.close()

    menu_path.write_text(
        json.dumps(
            {
                "children": [
                    {
                        "itemMasterId": 200,
                        "title": "Tea",
                        "itemPathKey": "1-200",
                        "displayAttribute": {"plu": "900002"},
                        "children": [],
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    counts = reconcile_workbook(
        source_path,
        menu_path,
        output_path,
        include_unmigrated=True,
    )

    assert counts[STATUS_MATCH] == 1
    result = load_workbook(output_path, data_only=True)
    mapping = result["Aloha > Qu Item Migration"]
    headers = {
        cell.value: cell.column for cell in mapping[8] if cell.value
    }
    assert mapping.cell(9, headers["Migrated to Qu?"]).value == "Yes"
    assert mapping.cell(9, headers["Qu Item ID"]).value == 200
    result.close()
