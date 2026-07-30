"""Print reconciliation-level PLU mismatches from the latest menu."""

from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
import re

from openpyxl import load_workbook

from src.matcher import format_path, is_combo_path
from src.menu_parser import flatten_menu, index_by_item_id
from src.models import MenuNode, ReconciliationStatus
from src.normalization import normalize_plu
from src.reconciliation import HEADER_ROW, MAPPING_SHEET

from .comparison import limited, parse_args, print_heading


@dataclass(frozen=True)
class PluMismatchRecord:
    aloha_plu: str
    aloha_name: str
    qu_item_id: int
    qu_item_name: str
    notes: str


def load_plu_mismatch_records(
    workbook_path: Path,
) -> list[PluMismatchRecord]:
    if not workbook_path.exists():
        raise FileNotFoundError(
            f"Reconciled workbook was not found: {workbook_path}. "
            "Run `python -m scripts.run_workflow` first."
        )

    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )
    try:
        if MAPPING_SHEET not in workbook.sheetnames:
            raise ValueError(
                f"Workbook does not contain {MAPPING_SHEET!r}."
            )

        sheet = workbook[MAPPING_SHEET]
        rows = sheet.iter_rows(
            min_row=HEADER_ROW,
            values_only=True,
        )
        headers = next(rows, None)
        if headers is None:
            raise ValueError("Mapping workbook has no header row.")

        columns = {
            str(value).strip(): index
            for index, value in enumerate(headers)
            if value is not None
        }
        required = {
            "Number",
            "Long name",
            "Qu Item ID",
            "Qu Item Name",
            "Reconciliation Status",
            "Reconciliation Notes",
        }
        missing = required - columns.keys()
        if missing:
            raise ValueError(
                f"Missing reconciliation columns: {sorted(missing)}"
            )

        records = []
        expected_status = (
            ReconciliationStatus.PLU_MISMATCH_REVIEW.value
        )
        for row in rows:
            status = str(
                row[columns["Reconciliation Status"]] or ""
            ).strip()
            if status != expected_status:
                continue

            item_id = row[columns["Qu Item ID"]]
            if item_id in (None, ""):
                continue

            records.append(
                PluMismatchRecord(
                    aloha_plu=normalize_plu(
                        row[columns["Number"]]
                    ),
                    aloha_name=str(
                        row[columns["Long name"]] or ""
                    ).strip(),
                    qu_item_id=int(item_id),
                    qu_item_name=str(
                        row[columns["Qu Item Name"]] or ""
                    ).strip(),
                    notes=str(
                        row[columns["Reconciliation Notes"]] or ""
                    ).strip(),
                )
            )
        return records
    finally:
        workbook.close()


def load_menu_nodes(
    menu_path: Path,
) -> tuple[list[MenuNode], dict[int, list[MenuNode]]]:
    if not menu_path.exists():
        raise FileNotFoundError(
            f"Cached menu was not found: {menu_path}. "
            "Run `python -m src.main` first."
        )

    with menu_path.open(encoding="utf-8-sig") as source:
        menu = json.load(source)
    nodes = [
        node
        for node in flatten_menu(menu)
        if not node.deleted
    ]
    return nodes, index_by_item_id(nodes)


def _name_tokens(value: str) -> set[str]:
    normalized = value.casefold().replace("shackburger", "burger")
    tokens = set(re.findall(r"[a-z0-9]+", normalized))
    return tokens - {"bbq", "combo", "shack", "style"}


def find_likely_descendants(
    record: PluMismatchRecord,
    existing_nodes: list[MenuNode],
    all_nodes: list[MenuNode],
) -> list[MenuNode]:
    """Find name-similar descendants for review without remapping."""
    expected_tokens = _name_tokens(record.aloha_name)
    if not expected_tokens:
        return []

    prefixes = tuple(
        f"{node.item_path_key}-"
        for node in existing_nodes
        if node.item_path_key
    )
    if not prefixes:
        return []

    scored: list[tuple[float, MenuNode]] = []
    seen: set[tuple[int | None, str | None]] = set()
    for node in all_nodes:
        if (
            node.item_master_id == record.qu_item_id
            or not node.item_path_key
            or not node.item_path_key.startswith(prefixes)
        ):
            continue

        tokens = _name_tokens(node.title)
        if not tokens:
            continue
        score = len(expected_tokens & tokens) / len(
            expected_tokens | tokens
        )
        if score < 0.6:
            continue

        identity = (node.item_master_id, node.item_path_key)
        if identity in seen:
            continue
        seen.add(identity)
        scored.append((score, node))

    scored.sort(
        key=lambda item: (
            -item[0],
            len(item[1].ancestors),
            item[1].title.casefold(),
        )
    )
    return [node for _, node in scored[:5]]


def _context_label(node: MenuNode) -> str:
    return "COMBO" if is_combo_path(node) else "REGULAR"


def print_record(
    record: PluMismatchRecord,
    existing_nodes: list[MenuNode],
    likely_nodes: list[MenuNode],
) -> None:
    actual_plus = sorted(
        {
            normalize_plu(node.plu)
            for node in existing_nodes
            if normalize_plu(node.plu)
        }
    )
    regular_count = sum(
        not is_combo_path(node) for node in existing_nodes
    )
    combo_count = len(existing_nodes) - regular_count

    print(
        f"\nAloha PLU={record.aloha_plu} | "
        f"Aloha Name={record.aloha_name}"
    )
    print(
        f"Existing QU ID={record.qu_item_id} | "
        f"QU Name={record.qu_item_name}"
    )
    print(
        f"Current PLUs={actual_plus or ['(blank)']} | "
        f"Regular Paths={regular_count} | "
        f"Combo Paths={combo_count}"
    )
    for node in existing_nodes:
        print(
            f"  [{_context_label(node)}] "
            f"{format_path(node)} | PLU={normalize_plu(node.plu) or '(blank)'}"
        )

    if likely_nodes:
        print("  Likely name matches (review only; not auto-mapped):")
        for node in likely_nodes:
            print(
                f"    [{_context_label(node)}] "
                f"QU ID={node.item_master_id} | "
                f"{format_path(node)} | "
                f"PLU={normalize_plu(node.plu) or '(blank)'}"
            )
    print(f"Reason: {record.notes}")


def main() -> None:
    args = parse_args(
        "Print reconciliation PLU mismatches from the latest menu."
    )
    records = load_plu_mismatch_records(args.workbook)
    all_nodes, menu_index = load_menu_nodes(args.menu)

    print(f"Workbook: {args.workbook}")
    print(f"Menu: {args.menu}")
    print_heading("PLU_MISMATCH_REVIEW", len(records))
    for record in limited(records, args.limit):
        existing_nodes = menu_index.get(record.qu_item_id, [])
        likely_nodes = find_likely_descendants(
            record,
            existing_nodes,
            all_nodes,
        )
        print_record(record, existing_nodes, likely_nodes)


if __name__ == "__main__":
    main()
