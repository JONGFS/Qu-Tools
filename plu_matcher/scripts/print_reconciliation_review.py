"""Print reconciliation rows that require manual review."""

import argparse
from pathlib import Path

from openpyxl import load_workbook

from src.models import SAFE_RECONCILIATION_STATUSES


DEFAULT_WORKBOOK = Path(
    "outputs/Aloha_Qu_Menu_Reconciled.xlsx"
)
SAFE_STATUS_VALUES = {
    status.value for status in SAFE_RECONCILIATION_STATUSES
}


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Print unsafe reconciliation rows for review."
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
    )
    args = parser.parse_args()

    if not args.workbook.exists():
        raise FileNotFoundError(
            f"Reconciled workbook was not found: {args.workbook}. "
            "Run `python -m scripts.reconcile_aloha_mappings` first."
        )

    workbook = load_workbook(
        args.workbook,
        read_only=True,
        data_only=True,
    )
    try:
        sheet = workbook["Aloha > Qu Item Migration"]
        rows = sheet.iter_rows(min_row=8, values_only=True)
        headers = next(rows)
        column = {
            str(value).strip(): index
            for index, value in enumerate(headers)
            if value is not None
        }

        review_rows = []
        for row in rows:
            status = str(
                row[column["Reconciliation Status"]] or ""
            ).strip()
            if not status or status in SAFE_STATUS_VALUES:
                continue
            review_rows.append(row)

        print(f"RECONCILIATION REVIEW: {len(review_rows)}")
        print("=" * 80)
        selected = (
            review_rows[: args.limit]
            if args.limit > 0
            else review_rows
        )
        for row in selected:
            print(
                f"PLU={row[column['Number']]} | "
                f"Aloha Name={row[column['Long name']]} | "
                f"Status={row[column['Reconciliation Status']]} | "
                f"Existing QU ID={row[column['Qu Item ID']]} | "
                f"Proposed QU ID="
                f"{row[column['Proposed Qu Item ID']]} | "
                f"Candidates={row[column['Candidate QU Items']]} | "
                f"Notes={row[column['Reconciliation Notes']]}"
            )
    finally:
        workbook.close()


if __name__ == "__main__":
    main()
