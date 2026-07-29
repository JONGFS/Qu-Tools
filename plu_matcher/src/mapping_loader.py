"""Load validated mappings from Excel."""
from pathlib import Path

from openpyxl import load_workbook

from .models import (
    ApprovedMapping,
    ReconciliationStatus,
    SAFE_RECONCILIATION_STATUSES,
)
from .normalization import normalize_plu


def load_mappings(workbook_path: Path) -> list[ApprovedMapping]:
    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )

    try:
        sheet = workbook["Aloha > Qu Item Migration"]

        rows = sheet.iter_rows(min_row=8, values_only=True)
        headers = next(rows, None)
        if headers is None:
            raise ValueError("Mapping workbook has no header row.")

        column = {
            str(header).strip(): index
            for index, header in enumerate(headers)
            if header is not None
        }
        required = {
            "Number",
            "Long name",
            "Migrated to Qu?",
            "Qu Item ID",
            "Qu Item Name",
        }
        missing = required - column.keys()
        if missing:
            raise ValueError(
                f"Missing required mapping columns: {sorted(missing)}"
            )

        status_column = column.get("Reconciliation Status")
        mappings: list[ApprovedMapping] = []

        for row_number, row in enumerate(rows, start=9):
            migrated = str(
                row[column["Migrated to Qu?"]] or ""
            ).strip().casefold()

            if status_column is not None:
                raw_status = str(row[status_column] or "").strip()
                try:
                    status = ReconciliationStatus(raw_status)
                except ValueError:
                    continue
                if status not in SAFE_RECONCILIATION_STATUSES:
                    continue
            elif migrated != "yes":
                continue

            aloha_plu = normalize_plu(row[column["Number"]])
            qu_item_id = row[column["Qu Item ID"]]
            if not aloha_plu:
                raise ValueError(
                    f"Safe mapping on row {row_number} has no Aloha PLU."
                )
            if qu_item_id in (None, ""):
                raise ValueError(
                    f"Safe mapping on row {row_number} has no QU item ID."
                )

            mappings.append(
                ApprovedMapping(
                    aloha_plu=aloha_plu,
                    aloha_name=str(
                        row[column["Long name"]] or ""
                    ).strip(),
                    qu_item_id=int(qu_item_id),
                    qu_name=str(
                        row[column["Qu Item Name"]] or ""
                    ).strip(),
                    old_item_path_key=None,
                )
            )
        return mappings
    finally:
        workbook.close()
