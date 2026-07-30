"""Generate portable CSV reports for reconciliation and path validation."""

from __future__ import annotations

from collections import Counter
import csv
from pathlib import Path

from openpyxl import load_workbook

from .models import (
    PathResolution,
    SAFE_RECONCILIATION_STATUSES,
)
from .reconciliation import (
    AUDIT_HEADERS,
    HEADER_ROW,
    MAPPING_SHEET,
)


PATH_REPORT_HEADERS = (
    "Aloha PLU",
    "Aloha Name",
    "QU Item ID",
    "Expected QU Name",
    "Current QU Name",
    "Old Path",
    "Current Path",
    "Ancestors",
    "Status",
    "Reason",
)

RECONCILIATION_REPORT_HEADERS = (
    "Number",
    "Long name",
    "Migrated to Qu?",
    "Qu Item ID",
    "Qu Item Name",
    *AUDIT_HEADERS,
)

SAFE_STATUS_VALUES = {
    status.value for status in SAFE_RECONCILIATION_STATUSES
}


def _write_rows(
    output_path: Path,
    headers: tuple[str, ...],
    rows: list[dict[str, object]],
) -> None:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as destination:
        writer = csv.DictWriter(
            destination,
            fieldnames=headers,
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(rows)


def write_csv_report(
    results: list[PathResolution],
    output_path: Path,
) -> None:
    """Write path-resolution results without discarding missing matches."""
    rows: list[dict[str, object]] = []
    for result in results:
        mapping = result.mapping
        current = result.current_node
        rows.append(
            {
                "Aloha PLU": mapping.aloha_plu,
                "Aloha Name": mapping.aloha_name,
                "QU Item ID": mapping.qu_item_id,
                "Expected QU Name": mapping.qu_name,
                "Current QU Name": current.title if current else "",
                "Old Path": mapping.old_item_path_key or "",
                "Current Path": (
                    current.item_path_key if current else ""
                ),
                "Ancestors": (
                    " > ".join(current.ancestors)
                    if current
                    else ""
                ),
                "Status": result.status.value,
                "Reason": result.reason,
            }
        )
    _write_rows(Path(output_path), PATH_REPORT_HEADERS, rows)


def export_reconciliation_reports(
    workbook_path: Path,
    all_csv_path: Path,
    review_csv_path: Path,
) -> tuple[Counter[str], int]:
    """Export all processed reconciliation rows and the unsafe subset.

    A processed row has a non-empty ``Reconciliation Status``. Statuses
    listed in ``SAFE_RECONCILIATION_STATUSES`` are omitted from the review
    report. Unknown non-empty statuses are deliberately treated as unsafe.
    """
    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )
    try:
        if MAPPING_SHEET not in workbook.sheetnames:
            raise ValueError(
                f"Workbook does not contain sheet {MAPPING_SHEET!r}."
            )

        sheet = workbook[MAPPING_SHEET]
        values = sheet.iter_rows(
            min_row=HEADER_ROW,
            values_only=True,
        )
        try:
            raw_headers = next(values)
        except StopIteration as error:
            raise ValueError(
                f"Workbook has no header row at row {HEADER_ROW}."
            ) from error

        column = {
            str(value).strip(): index
            for index, value in enumerate(raw_headers)
            if value is not None and str(value).strip()
        }
        missing = set(RECONCILIATION_REPORT_HEADERS) - column.keys()
        if missing:
            raise ValueError(
                "Missing required reconciliation report columns: "
                f"{sorted(missing)}"
            )

        all_rows: list[dict[str, object]] = []
        review_rows: list[dict[str, object]] = []
        counts: Counter[str] = Counter()

        for values_row in values:
            status_value = values_row[
                column["Reconciliation Status"]
            ]
            status = str(status_value or "").strip()
            if not status:
                continue

            row = {
                header: values_row[column[header]]
                for header in RECONCILIATION_REPORT_HEADERS
            }
            # Aloha PLUs are identifiers and should never be serialized as
            # floating-point values in a portable report.
            number = row["Number"]
            if isinstance(number, float) and number.is_integer():
                number = int(number)
            row["Number"] = (
                ""
                if number is None
                else str(number).strip()
            )
            row["Reconciliation Status"] = status

            all_rows.append(row)
            counts[status] += 1
            if status not in SAFE_STATUS_VALUES:
                review_rows.append(row)
    finally:
        workbook.close()

    _write_rows(
        Path(all_csv_path),
        RECONCILIATION_REPORT_HEADERS,
        all_rows,
    )
    _write_rows(
        Path(review_csv_path),
        RECONCILIATION_REPORT_HEADERS,
        review_rows,
    )
    return counts, len(review_rows)
