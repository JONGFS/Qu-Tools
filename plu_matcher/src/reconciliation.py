"""Safely reconcile Aloha PLUs with the current QU menu."""

from __future__ import annotations

from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from .menu_parser import flatten_menu
from .models import (
    MenuNode,
    ReconciliationStatus,
)
from .normalization import normalize_plu


MAPPING_SHEET = "Aloha > Qu Item Migration"
HEADER_ROW = 8

STATUS_MATCH = ReconciliationStatus.MATCH
STATUS_MULTIPLE_PATHS = ReconciliationStatus.MATCH_MULTIPLE_PATHS
STATUS_ID_UPDATED = ReconciliationStatus.PLU_MATCH_ID_UPDATED
STATUS_AMBIGUOUS = ReconciliationStatus.AMBIGUOUS_QU_ITEM
STATUS_CONFLICT = ReconciliationStatus.EXISTING_MAPPING_CONFLICT
STATUS_MISMATCH = ReconciliationStatus.PLU_MISMATCH_REVIEW
STATUS_NOT_FOUND = ReconciliationStatus.NOT_FOUND_IN_QU
STATUS_NOT_IN_CONTEXT = (
    ReconciliationStatus.NOT_FOUND_IN_CURRENT_CONTEXT
)

AUDIT_HEADERS = (
    "Reconciliation Status",
    "Previous Qu Item ID",
    "Previous Qu Item Name",
    "Proposed Qu Item ID",
    "Proposed Qu Item Name",
    "QU Path Count",
    "Candidate QU Items",
    "Reconciliation Notes",
)


@dataclass(frozen=True)
class MenuIndexes:
    by_plu: dict[str, list[MenuNode]]
    by_item_id: dict[int, list[MenuNode]]


@dataclass(frozen=True)
class ReconciliationDecision:
    status: ReconciliationStatus
    proposed_item_id: int | None
    proposed_item_name: str | None
    path_count: int
    candidates: str
    notes: str
    apply_mapping: bool = False


def build_menu_indexes(menu: dict) -> MenuIndexes:
    by_plu: dict[str, list[MenuNode]] = defaultdict(list)
    by_item_id: dict[int, list[MenuNode]] = defaultdict(list)

    for node in flatten_menu(menu):
        if node.deleted or node.item_master_id is None:
            continue

        by_item_id[node.item_master_id].append(node)
        plu = normalize_plu(node.plu)
        if plu:
            by_plu[plu].append(node)

    return MenuIndexes(
        by_plu=dict(by_plu),
        by_item_id=dict(by_item_id),
    )


def _preferred_title(nodes: list[MenuNode]) -> str:
    titles = [
        node.title.strip()
        for node in nodes
        if node.title and node.title.strip()
    ]
    if not titles:
        return ""

    counts = Counter(titles)
    return min(
        counts,
        key=lambda title: (-counts[title], title.casefold()),
    )


def _candidate_description(
    nodes_by_id: dict[int, list[MenuNode]],
) -> str:
    descriptions = []
    for item_id in sorted(nodes_by_id):
        nodes = nodes_by_id[item_id]
        titles = sorted(
            {
                node.title.strip()
                for node in nodes
                if node.title and node.title.strip()
            },
            key=str.casefold,
        )
        title_text = " / ".join(titles) or "(untitled)"
        descriptions.append(
            f"{item_id}: {title_text} ({len(nodes)} paths)"
        )
    return "; ".join(descriptions)


def decide_reconciliation(
    aloha_plu: str,
    existing_item_id: int | None,
    existing_item_name: str,
    indexes: MenuIndexes,
) -> ReconciliationDecision:
    exact_nodes = indexes.by_plu.get(aloha_plu, [])
    exact_by_id: dict[int, list[MenuNode]] = defaultdict(list)
    for node in exact_nodes:
        if node.item_master_id is not None:
            exact_by_id[node.item_master_id].append(node)

    if not exact_by_id:
        if existing_item_id is None:
            return ReconciliationDecision(
                status=STATUS_NOT_FOUND,
                proposed_item_id=None,
                proposed_item_name=None,
                path_count=0,
                candidates="",
                notes="Exact Aloha PLU was not found in the current QU menu.",
            )

        existing_nodes = indexes.by_item_id.get(existing_item_id, [])
        if not existing_nodes:
            return ReconciliationDecision(
                status=STATUS_NOT_IN_CONTEXT,
                proposed_item_id=None,
                proposed_item_name=None,
                path_count=0,
                candidates="",
                notes=(
                    "Existing QU mapping was preserved, but its item ID "
                    "was not found in the current menu context."
                ),
            )

        actual_plus = sorted(
            {
                normalize_plu(node.plu)
                for node in existing_nodes
                if normalize_plu(node.plu)
            }
        )
        return ReconciliationDecision(
            status=STATUS_MISMATCH,
            proposed_item_id=None,
            proposed_item_name=None,
            path_count=len(existing_nodes),
            candidates=_candidate_description(
                {existing_item_id: existing_nodes}
            ),
            notes=(
                f"Existing QU mapping was preserved. Expected PLU "
                f"{aloha_plu}; current PLUs for item {existing_item_id}: "
                f"{actual_plus or ['(blank)']}."
            ),
        )

    candidate_text = _candidate_description(dict(exact_by_id))
    total_paths = sum(len(nodes) for nodes in exact_by_id.values())

    if len(exact_by_id) > 1:
        return ReconciliationDecision(
            status=STATUS_AMBIGUOUS,
            proposed_item_id=None,
            proposed_item_name=None,
            path_count=total_paths,
            candidates=candidate_text,
            notes=(
                "Exact PLU belongs to multiple QU item IDs. Existing "
                "mapping was preserved and manual review is required."
            ),
        )

    proposed_item_id, proposed_nodes = next(iter(exact_by_id.items()))
    proposed_name = _preferred_title(proposed_nodes)

    if (
        existing_item_id is not None
        and existing_item_id != proposed_item_id
    ):
        return ReconciliationDecision(
            status=STATUS_ID_UPDATED,
            proposed_item_id=proposed_item_id,
            proposed_item_name=proposed_name,
            path_count=len(proposed_nodes),
            candidates=candidate_text,
            notes=(
                f"Exact PLU {aloha_plu} uniquely identifies QU item "
                f"{proposed_item_id}; previous QU item "
                f"{existing_item_id} "
                f"({existing_item_name or 'unnamed'}) was replaced."
            ),
            apply_mapping=True,
        )

    status = (
        STATUS_MATCH
        if len(proposed_nodes) == 1
        else STATUS_MULTIPLE_PATHS
    )
    notes = (
        "Exact PLU matched one unique QU item ID."
        if status == STATUS_MATCH
        else (
            "Exact PLU matched one QU item ID across multiple paths; "
            "item identity is safe, but path selection remains separate."
        )
    )
    return ReconciliationDecision(
        status=status,
        proposed_item_id=proposed_item_id,
        proposed_item_name=proposed_name,
        path_count=len(proposed_nodes),
        candidates=candidate_text,
        notes=notes,
        apply_mapping=True,
    )


def _header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[HEADER_ROW]
        if cell.value is not None
    }


def _add_audit_headers(sheet) -> dict[str, int]:
    columns = _header_map(sheet)
    start_column = sheet.max_column + 1
    style_source = sheet.cell(HEADER_ROW, min(8, sheet.max_column))

    for offset, header in enumerate(AUDIT_HEADERS):
        column = columns.get(header, start_column + offset)
        cell = sheet.cell(HEADER_ROW, column, header)
        if style_source.has_style:
            cell._style = copy(style_source._style)
        cell.font = copy(style_source.font)
        cell.fill = copy(style_source.fill)
        cell.border = copy(style_source.border)
        cell.alignment = copy(style_source.alignment)
        cell.number_format = style_source.number_format
        columns[header] = column

    widths = {
        "Reconciliation Status": 28,
        "Previous Qu Item ID": 20,
        "Previous Qu Item Name": 32,
        "Proposed Qu Item ID": 20,
        "Proposed Qu Item Name": 32,
        "QU Path Count": 15,
        "Candidate QU Items": 60,
        "Reconciliation Notes": 70,
    }
    for header, width in widths.items():
        letter = sheet.cell(HEADER_ROW, columns[header]).column_letter
        sheet.column_dimensions[letter].width = width

    return columns


def reconcile_workbook(
    source_workbook: Path,
    menu_path: Path,
    output_workbook: Path,
    *,
    include_unmigrated: bool = False,
    provenance: dict | None = None,
) -> Counter:
    if source_workbook.resolve() == output_workbook.resolve():
        raise ValueError("Output workbook must not overwrite the source.")

    with menu_path.open(encoding="utf-8-sig") as source:
        menu = json.load(source)
    indexes = build_menu_indexes(menu)

    workbook = load_workbook(source_workbook)
    value_workbook = load_workbook(
        source_workbook,
        data_only=True,
    )
    if MAPPING_SHEET not in workbook.sheetnames:
        workbook.close()
        value_workbook.close()
        raise ValueError(
            f"Workbook does not contain sheet {MAPPING_SHEET!r}."
        )

    sheet = workbook[MAPPING_SHEET]
    value_sheet = value_workbook[MAPPING_SHEET]
    columns = _add_audit_headers(sheet)
    required = {
        "Number",
        "Long name",
        "Migrated to Qu?",
        "Qu Item ID",
        "Qu Item Name",
    }
    missing = required - columns.keys()
    if missing:
        workbook.close()
        value_workbook.close()
        raise ValueError(
            f"Missing required workbook columns: {sorted(missing)}"
        )

    counts: Counter = Counter()
    for row_number in range(HEADER_ROW + 1, sheet.max_row + 1):
        aloha_plu = normalize_plu(
            sheet.cell(row_number, columns["Number"]).value
        )
        if not aloha_plu:
            continue

        migrated = str(
            value_sheet.cell(
                row_number,
                columns["Migrated to Qu?"],
            ).value
            or ""
        ).strip().casefold()
        if migrated != "yes" and not include_unmigrated:
            for header in AUDIT_HEADERS:
                sheet.cell(row_number, columns[header]).value = None
            continue

        existing_id_value = value_sheet.cell(
            row_number,
            columns["Qu Item ID"],
        ).value
        existing_item_id = (
            int(existing_id_value)
            if existing_id_value not in (None, "")
            else None
        )
        existing_name = str(
            value_sheet.cell(
                row_number,
                columns["Qu Item Name"],
            ).value
            or ""
        ).strip()

        # The reconciled workbook is a static, auditable snapshot. Materialize
        # cached lookup results so downstream Python reads do not depend on
        # Excel recalculating XLOOKUP formulas first.
        item_id_cell = sheet.cell(
            row_number,
            columns["Qu Item ID"],
        )
        item_name_cell = sheet.cell(
            row_number,
            columns["Qu Item Name"],
        )
        item_id_cell.value = existing_item_id
        item_name_cell.value = existing_name or None

        decision = decide_reconciliation(
            aloha_plu=aloha_plu,
            existing_item_id=existing_item_id,
            existing_item_name=existing_name,
            indexes=indexes,
        )
        counts[decision.status] += 1

        audit_values = {
            "Reconciliation Status": decision.status,
            "Previous Qu Item ID": existing_item_id,
            "Previous Qu Item Name": existing_name or None,
            "Proposed Qu Item ID": decision.proposed_item_id,
            "Proposed Qu Item Name": decision.proposed_item_name,
            "QU Path Count": decision.path_count,
            "Candidate QU Items": decision.candidates,
            "Reconciliation Notes": decision.notes,
        }
        for header, value in audit_values.items():
            sheet.cell(row_number, columns[header], value)

        if decision.apply_mapping:
            sheet.cell(
                row_number,
                columns["Migrated to Qu?"],
                "Yes",
            )
            item_id_cell.value = decision.proposed_item_id
            item_name_cell.value = decision.proposed_item_name

    summary_name = "Reconciliation Summary"
    if summary_name in workbook.sheetnames:
        workbook.remove(workbook[summary_name])
    summary = workbook.create_sheet(summary_name, 0)
    summary.append(["Aloha to QU Reconciliation", None])
    menu_context = menu.get("context") or {}
    provenance = provenance or {}
    summary.append(["Source workbook", source_workbook.name])
    summary.append(["Menu JSON", str(menu_path)])
    summary.append(["Menu snapshot ID", menu.get("menuSnapshotId")])
    summary.append(
        [
            "Menu generation time",
            provenance.get("generation_time"),
        ]
    )
    summary.append(
        [
            "Location ID",
            menu_context.get("locationId")
            or provenance.get("location_id"),
        ]
    )
    summary.append(
        [
            "Order channel ID",
            menu_context.get("orderChannelId")
            or provenance.get("order_channel_id"),
        ]
    )
    summary.append(
        [
            "Order type ID",
            menu_context.get("orderTypeId")
            or provenance.get("order_type_id"),
        ]
    )
    summary.append(
        [
            "Cache last checked",
            provenance.get("last_checked_at"),
        ]
    )
    summary.append(["Status", "Count"])
    for status, count in sorted(counts.items()):
        summary.append([status, count])
    summary.append(["TOTAL", sum(counts.values())])

    summary["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary.merge_cells("A1:B1")
    status_header_row = summary.max_row - len(counts) - 1
    for cell in summary[status_header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 55
    summary.freeze_panes = f"A{status_header_row + 1}"

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_workbook.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_workbook)
    workbook.close()
    value_workbook.close()
    return counts
