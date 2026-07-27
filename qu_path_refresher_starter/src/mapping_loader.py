"""Load validated mappings from Excel."""
from pathlib import Path
from .models import ApprovedMapping
from openpyxl import load_workbook

def load_mappings(workbook_path: Path) -> list[ApprovedMapping]:
    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True
    )
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)


def normalize_plu(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value,float) and value.is_integer():
        return str(int(value))
    return str(value).strip()

def load_mappings(workbook_path: Path) -> list[ApprovedMapping]:
    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True
    )

    sheet = workbook["Aloha > Qu Item Migration"]

    rows = sheet.iter_rows(min_row=8,values_only=True)
    headers = next(rows)

    column = {
        str(header).strip(): index 
        for index, header in enumerate(headers)
        if header is not None
    }

    mappings: list[ApprovedMapping] = []

    for row in rows:
    
        migrated = str(row[column["Migrated to Qu?"]] or "").strip()

        if migrated.casefold() != "yes":
            continue

        mappings.append( 
            ApprovedMapping(
                aloha_plu=normalize_plu(row[column["Number"]]),
                aloha_name=str(row[column["Long name"]] or "").strip(),
                qu_item_id=int(row[column["Qu Item ID"]]),
                qu_name=str(row[column["Qu Item Name"]] or "").strip(),
                old_item_path_key=None

            )
        )
    workbook.close()
    return mappings